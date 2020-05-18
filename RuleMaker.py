"""
Reads ObjectsExt.csv,
creates for MD, like
c1: (u, v, f), (u, v, f) ... U
...
"""

import numpy as np
import json
import functions

np.seterr(all="ignore")


def turgunlik_ishchi(feature_col, labels, explain_text):

    ORDERED_COLUMN = 1

    m = len(feature_col)
    matrix = np.array(feature_col).reshape((m, 1))
    id_col = np.arange(m).reshape((m, 1))
    matrix = np.hstack((id_col, matrix, labels.reshape((m, 1))))

    # print("Unsorted", matrix)
    matrix = functions.sorting_by_col(matrix, ORDERED_COLUMN)
    # print("Sorted", matrix)
    labels = matrix[:, 2].copy()
    class_names, class_members = np.unique(labels, return_counts=True)
    indexes = functions.get_uv(matrix, col_n=ORDERED_COLUMN)
    # indexes = [int(x) for x in matritsa[:, 0]]
    # print(indexes)

    # intervals = functions.finding_intervals(labels, indexes, class_members)
    intervals = functions.finding_intervals_with_tegishlilik_f(labels, indexes, class_members)

    # intervals.sort(key=lambda x: x[2], reverse=True)

    # intervallar = 0     # 'intervallar'
    # tegishlilik_f = 0   # 'har bir intervallning tegishlilik funksiyasi qiymati'
    # m = 0               # Obyektlar soni

    intervals2 = []
    summa = 0
    for interval in intervals:
        if interval[4] > .5: summa += interval[4] * (interval[1]-interval[0])
        else: summa += (1 - interval[4]) * (interval[1] - interval[0])
        # best_interval = (u, v, summa, int(dominator), tegishlilik_f, nu1, nu2)
        intervals2.append((matrix[interval[0], 1], matrix[interval[1], 1], # u-val, v_val
                      interval[3], interval[4],   # dominator, teg_f
                      ))


    javob = summa / m
    javob = (javob, len(intervals), explain_text)
    #print(f"{explain_text} TURG'UNLIGI: {javob:1.2f} :: intervallar: {intervals}")

    return intervals2, javob


project_path = r"d:\_NUU\2020\batonMD"
matritsa = np.genfromtxt(project_path + r"\init_data\giper\ObjectsExt.csv", delimiter=',')
labels = np.genfromtxt(project_path + r"\init_data\giper\Target.csv", delimiter=",")
with open(project_path + r"\init_data\giper\feature_names.csv") as f:
    feature_names = f.read().split("\n")
# print(feature_names)

shape = matritsa.shape
# feature_sorting = []
# sort_turgunlik = sorted(feature_sorting, key=lambda L:L[0], reverse=True)

data = {"info": "info about datastruct", "features_count": shape[1]}
data["intervals_structure"] = """Интервалы объеденены в 1 массив вида
[(<интерва1>), (<интервал2>),..., (<интервалq>)]. Каждый интервал содержит инфо вида
(<левая_граница>, <правая_граница>, <доминатор>, <зн_функ_принадлежности>)"""

for col in range(shape[1]):
    intervals, tur = turgunlik_ishchi(matritsa[:, col], labels, "sth")
    # print(tur)
    # print(intervals)

    data[col] = {
        "feature_name": feature_names[col],
        "stability": tur[0],
        "intervals_count": tur[1],
        "intervals": intervals.copy(),
    }

print(data.keys())
for d in range(data["features_count"]):
    print(data[d])

with open("out_data/fuzzy_logic.json", "w", encoding="utf-8") as file:
    json.dump(data, file)

# EXPORT DATA INTO xlsx file
print('\n\n########export data###################\n\n')
from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "Table 3"
ws.sheet_properties.tabColor = "008000"
# ws.cell(row=5, column=5, value='asdasdas')
# ws.cell(row=1, column=6, value=123)
# ws.cell(row=1, column=1, value=12.3)

ws.cell(row=1, column=1, value="Alomat nomi")
ws.cell(row=1, column=2, value="Intervallar soni")
ws.cell(row=1, column=3, value="Intervallar, tegislilik funsiyalari va dominator")
ws.cell(row=1, column=4, value="U")

row = 2
for i in range(data['features_count']):

    ws.cell(row=row, column=1, value=data[i]['feature_name'])
    ws.cell(row=row, column=2, value=data[i]['intervals_count'])

    s = ""
    for interval in data[i]['intervals']:
        s += f"[{interval[0]:.2f}..{interval[1]:.2f}]({interval[3]:.2f})K{interval[2]}, "
    ws.cell(row=row, column=3, value=s)

    ws.cell(row=row, column=4, value=f"\t{data[i]['stability']:.2f}")
    row += 1

wb.save('out_data/table3.xlsx')
